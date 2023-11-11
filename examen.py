import tkinter as tk
from tkinter import messagebox
import openpyxl

class MegaAutosApp:
    def _init_(self, master):
        self.master = master
        self.master.title("Mega Autos App")
        
        
        self.codigo_label = tk.Label(master, text="Código:")
        self.codigo_label.grid(row=0, column=0)
        self.codigo_entry = tk.Entry(master)
        self.codigo_entry.grid(row=0, column=1)

        self.marca_label = tk.Label(master, text="Marca:")
        self.marca_label.grid(row=1, column=0)
        self.marca_entry = tk.Entry(master)
        self.marca_entry.grid(row=1, column=1)

        self.modelo_label = tk.Label(master, text="Modelo:")
        self.modelo_label.grid(row=2, column=0)
        self.modelo_entry = tk.Entry(master)
        self.modelo_entry.grid(row=2, column=1)

        self.precio_label = tk.Label(master, text="Precio:")
        self.precio_label.grid(row=3, column=0)
        self.precio_entry = tk.Entry(master)
        self.precio_entry.grid(row=3, column=1)

        self.kilometraje_label = tk.Label(master, text="Kilometraje:")
        self.kilometraje_label.grid(row=4, column=0)
        self.kilometraje_entry = tk.Entry(master)
        self.kilometraje_entry.grid(row=4, column=1)

        self.guardar_button = tk.Button(master, text="Guardar Vehículo", command=self.guardar_vehiculo)
        self.guardar_button.grid(row=5, column=0, columnspan=2)

        self.listar_button = tk.Button(master, text="Listar Vehículos", command=self.listar_vehiculos)
        self.listar_button.grid(row=6, column=0, columnspan=2)

    def guardar_vehiculo(self):
        
        codigo = self.codigo_entry.get()
        marca = self.marca_entry.get()
        modelo = self.modelo_entry.get()
        precio = float(self.precio_entry.get())
        kilometraje = int(self.kilometraje_entry.get())

       
        self.guardar_en_excel(codigo, marca, modelo, precio, kilometraje)

       
        self.codigo_entry.delete(0, tk.END)
        self.marca_entry.delete(0, tk.END)
        self.modelo_entry.delete(0, tk.END)
        self.precio_entry.delete(0, tk.END)
        self.kilometraje_entry.delete(0, tk.END)

    def guardar_en_excel(self, codigo, marca, modelo, precio, kilometraje):
        
        try:
            workbook = openpyxl.load_workbook('vehiculos.xlsx')
        except FileNotFoundError:
          
            workbook = openpyxl.Workbook()

        
        sheet = workbook['Listado'] if 'Listado' in workbook.sheetnames else workbook.create_sheet('Listado')

        
        nueva_fila = [codigo, marca, modelo, precio, kilometraje]
        sheet.append(nueva_fila)

        
        workbook.save('vehiculos.xlsx')
        messagebox.showinfo("Éxito", "Vehículo guardado exitosamente.")

    def listar_vehiculos(self):
        
        try:
            workbook = openpyxl.load_workbook('vehiculos.xlsx')
        except FileNotFoundError:
            messagebox.showwarning("Advertencia", "No hay vehículos para mostrar.")
            return

        
        sheet = workbook['Listado']

        
        datos = []
        for row in sheet.iter_rows(values_only=True):
            datos.append(row)

        if datos:
            mensaje = "\n".join([" | ".join(map(str, fila)) for fila in datos])
            messagebox.showinfo("Listado de Vehículos", mensaje)
        else:
            messagebox.showwarning("Advertencia", "No hay vehículos para mostrar.")

root = tk.Tk()
app = MegaAutosApp(root)
root.mainloop()